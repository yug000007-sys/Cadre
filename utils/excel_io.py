"""
Excel I/O — read and append rows to cadre_quotes.xlsx
"""

import os
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

from utils.extractor import COLUMNS

XLSX_PATH = os.environ.get("OUTPUT_XLSX", "data/cadre_quotes.xlsx")

HEADER_COLOR = "1F4E79"
ALT_ROW_COLOR = "EBF3FB"


def get_or_create_workbook(path: str = XLSX_PATH):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quotes"
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
    return wb, ws


def append_rows(rows: list[dict], path: str = XLSX_PATH) -> int:
    wb, ws = get_or_create_workbook(path)
    start_row = ws.max_row + 1
    for row_idx, row in enumerate(rows, start=start_row):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = row.get(col_name)
            # Convert date objects to string for Excel compatibility
            if isinstance(val, (date, datetime)):
                val = val.strftime("%m/%d/%Y") if isinstance(val, date) else val
            ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    wb.save(path)
    return len(rows)


def load_as_dataframe(path: str = XLSX_PATH) -> pd.DataFrame:
    if not os.path.exists(path):
        return pd.DataFrame(columns=COLUMNS)
    try:
        df = pd.read_excel(path, dtype=str)
        df = df.where(df.notna(), other=None)
        return df
    except Exception:
        return pd.DataFrame(columns=COLUMNS)


def quote_exists(quote_number: str, path: str = XLSX_PATH) -> bool:
    df = load_as_dataframe(path)
    if df.empty or "QuoteNumber" not in df.columns:
        return False
    return str(quote_number) in df["QuoteNumber"].astype(str).values


def get_stats(path: str = XLSX_PATH) -> dict:
    df = load_as_dataframe(path)
    if df.empty:
        return {"total_rows": 0, "unique_quotes": 0, "unique_customers": 0, "total_sales": 0.0}
    return {
        "total_rows":       len(df),
        "unique_quotes":    df["QuoteNumber"].nunique() if "QuoteNumber" in df.columns else 0,
        "unique_customers": df["Company"].nunique() if "Company" in df.columns else 0,
        "total_sales":      pd.to_numeric(df.get("TotalSales", pd.Series(dtype=float)), errors="coerce").sum(),
    }
